import openpyxl

def get_example_list(excelFile, sheetName = None, headerMatcher = None, headerUnmatcher = None):
    wb = openpyxl.load_workbook(excelFile)    
    # Define variable to read sheet
    if sheetName == None:
        ws = wb[wb.sheetnames[0]]
    else:
        if sheetName in wb.sheetnames:
            ws = wb[sheetName]
        else:
            raise Exception(sheetName + " Sheet is not found.")        
    # Iterate the loop to read the cell values
    IsFound = False
    parameterRow = 0
    parameterCol = 0
    for row in range(1, ws.max_row):
        for col in range(2, ws.max_column):
            if "Parameter Name" in str(ws.cell(row, col).value) :
                parameterRow = row
                parameterCol = col
                IsFound = True
                break
        if IsFound == True:
            break
    if IsFound == False:
        raise Exception("Paramter Name grid is not found.")
    # print(parameterRow)
    # print(parameterCol)
    parameterNames = "HeaderName"
    for row in range(parameterRow + 1, ws.max_row +1):
        if ws.cell(row, parameterCol).value != None and ws.cell(row, parameterCol).value != "NA" :
            parameterNames = parameterNames + ", " + str((ws.cell(row, parameterCol).value))

    print("The parameter names for test method is " + parameterNames)
    
    testcaseSetList = []
    for col in range(parameterCol+1, ws.max_column +1):
        if headerMatcher != None and headerMatcher not in str(ws.cell(parameterRow, col).value) :
            continue
        if headerUnmatcher != None and headerUnmatcher in str(ws.cell(parameterRow, col).value) :
            continue
        testcaseSet = []
        for row in range(parameterRow, ws.max_row +1):
            if ws.cell(row, parameterCol).value != None and ws.cell(row, parameterCol).value != "NA" :
                testcaseSet.append(ws.cell(row, col).value)
        testcaseSetList.append(testcaseSet)

    return testcaseSetList